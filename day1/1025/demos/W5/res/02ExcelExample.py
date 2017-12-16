import docx
import win32com
import win32com.client
from openpyxl.reader.excel import load_workbook


def writeExcel():

    # 调用系统接口打开Excel
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False#程序窗口不可见

    # 新建工作表并定位到当前sheet页
    wb = excel.Workbooks.Add()
    currentSheet = wb.ActiveSheet

    # 对指定行列的单元格赋值
    for row in range(1, 5):
        for col in range(1, 10):
            if row % 2 == 0 or col % 2 == 0:
                currentSheet.Cells(row, col).value = "Hello"
            else:
                currentSheet.Cells(row, col).value = "拍森"

    # 保存内存数据到文件，关闭IO流
    wb.SaveAs("C:\\Users\\idea\\Desktop\\悯码农2.xlsx")
    wb.Close(True)

    # 关闭Excel
    excel.Application.Quit()


def readExcel():
    # 载入指定路径的Excel文件到内存
    wb = load_workbook(r"C:\Users\idea\Desktop\悯码农2.xlsx")

    # 拿到文件中的所有工作表
    sheetNames = wb.get_sheet_names()
    print("wb.get_sheet_names=", sheetNames)
    # 打印第一个工作表的信息
    sheet = wb.get_sheet_by_name(sheetNames[0])

    print("sheet.title=", sheet.title)
    print("sheet.max_row=", sheet.max_row)
    print("sheet.max_column=", sheet.max_column)

    # 根据行列索引遍历每一个单元格内容
    for row in range(1, sheet.max_row + 1):
        line = ""
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=row, column=col).value
            if val == None:
                val = "——"
            line += str(val) + " "
        print(row, ":", line)


if __name__ == '__main__':
    # writeExcel()
    readExcel()
    pass
